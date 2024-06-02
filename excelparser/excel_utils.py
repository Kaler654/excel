import openpyxl


def parse_excel_file(file_path):
    workbook = openpyxl.load_workbook(file_path)
    data = {}

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        sheet_data = []
        for row in worksheet.iter_rows(values_only=True):
            parsed_row = []
            for value in row:
                if value is None:
                    continue
                if isinstance(value, (int, float)):  # Проверяем, является ли value числом
                    parsed_row.append(round(value))
                else:
                    parsed_row.append(value)
            sheet_data.append(parsed_row)
        data[sheet] = sheet_data

    return data

# def parse_excel_file(file_path):
#     workbook = openpyxl.load_workbook(file_path)
#     data = {}
#
#     for sheet in workbook.sheetnames:
#         worksheet = workbook[sheet]
#         sheet_data = []
#         for row in worksheet.iter_rows(values_only=True):
#             row = [value for value in row if value is not None]
#             print(row)
#             sheet_data.append(row)
#         data[sheet] = sheet_data
#
#     return data